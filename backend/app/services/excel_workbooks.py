import base64
from io import BytesIO
import hashlib
import hmac
import json
import secrets
import time
from urllib.parse import quote, urlencode

import httpx
from openpyxl import Workbook, load_workbook

from app.config import get_settings
from app.dependencies import get_supabase
from app.utils.logger import get_logger

logger = get_logger(__name__)

LEAD_HEADERS = [
    "ID",
    "Created At",
    "Patient Name",
    "Phone",
    "Email",
    "Reason",
    "Preferred Time",
    "Status",
]

AVAILABILITY_HEADERS = ["Date", "Time", "Status", "Patient Name", "Lead ID"]

MICROSOFT_CONNECT_SCOPES = [
    "offline_access",
    "openid",
    "profile",
    "email",
    "User.Read",
    "Files.ReadWrite",
]


def microsoft_excel_quick_connect_available() -> bool:
    settings = get_settings()
    return settings.microsoft_oauth_configured


def _microsoft_authority() -> str:
    settings = get_settings()
    tenant = (settings.microsoft_oauth_tenant or "common").strip() or "common"
    return f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0"


def build_microsoft_connect_url(redirect_uri: str, state: str) -> str:
    settings = get_settings()
    query = urlencode(
        {
            "client_id": settings.microsoft_oauth_client_id,
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "response_mode": "query",
            "scope": " ".join(MICROSOFT_CONNECT_SCOPES),
            "prompt": "select_account",
            "state": state,
        }
    )
    return f"{_microsoft_authority()}/authorize?{query}"


def sign_microsoft_connect_state(payload: dict) -> str:
    settings = get_settings()
    secret = (settings.admin_secret or settings.supabase_service_key).encode("utf-8")
    envelope = {
        **payload,
        "nonce": secrets.token_urlsafe(12),
        "ts": int(time.time()),
    }
    raw = json.dumps(envelope, separators=(",", ":"), sort_keys=True).encode("utf-8")
    signature = hmac.new(secret, raw, hashlib.sha256).hexdigest()
    return base64.urlsafe_b64encode(raw).decode("utf-8").rstrip("=") + "." + signature


def verify_microsoft_connect_state(state: str, max_age_seconds: int = 900) -> dict:
    settings = get_settings()
    secret = (settings.admin_secret or settings.supabase_service_key).encode("utf-8")

    try:
        encoded, signature = state.split(".", 1)
        padded = encoded + "=" * (-len(encoded) % 4)
        raw = base64.urlsafe_b64decode(padded.encode("utf-8"))
        expected = hmac.new(secret, raw, hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected):
            raise ValueError("Invalid signature")
        payload = json.loads(raw.decode("utf-8"))
    except Exception as exc:
        raise ValueError("Invalid Microsoft connect state") from exc

    ts = int(payload.get("ts") or 0)
    if not ts or time.time() - ts > max_age_seconds:
        raise ValueError("Microsoft connect state expired")

    return payload


async def exchange_microsoft_connect_code(code: str, redirect_uri: str) -> dict:
    settings = get_settings()
    async with httpx.AsyncClient(timeout=20.0) as client:
        response = await client.post(
            f"{_microsoft_authority()}/token",
            data={
                "client_id": settings.microsoft_oauth_client_id,
                "client_secret": settings.microsoft_oauth_client_secret,
                "grant_type": "authorization_code",
                "code": code,
                "redirect_uri": redirect_uri,
                "scope": " ".join(MICROSOFT_CONNECT_SCOPES),
            },
            headers={"Accept": "application/json"},
        )

    payload = response.json()
    if response.status_code >= 400:
        detail = payload.get("error_description") or payload.get("error") or "Microsoft token exchange failed"
        raise RuntimeError(detail)
    if not str(payload.get("access_token") or "").strip():
        raise RuntimeError("Microsoft token exchange did not return an access token")
    if not str(payload.get("refresh_token") or "").strip():
        raise RuntimeError("Microsoft token exchange did not return a refresh token")
    return payload


async def refresh_microsoft_access_token(refresh_token: str) -> dict:
    settings = get_settings()
    async with httpx.AsyncClient(timeout=20.0) as client:
        response = await client.post(
            f"{_microsoft_authority()}/token",
            data={
                "client_id": settings.microsoft_oauth_client_id,
                "client_secret": settings.microsoft_oauth_client_secret,
                "grant_type": "refresh_token",
                "refresh_token": refresh_token,
                "scope": " ".join(MICROSOFT_CONNECT_SCOPES),
            },
            headers={"Accept": "application/json"},
        )

    payload = response.json()
    if response.status_code >= 400:
        detail = payload.get("error_description") or payload.get("error") or "Microsoft token refresh failed"
        raise RuntimeError(detail)
    return payload


async def _graph_request(
    method: str,
    path: str,
    access_token: str,
    *,
    content: bytes | None = None,
    json_payload: dict | None = None,
) -> httpx.Response:
    headers = {"Authorization": f"Bearer {access_token}"}
    if json_payload is not None:
        headers["Content-Type"] = "application/json"
    async with httpx.AsyncClient(timeout=30.0) as client:
        return await client.request(
            method,
            f"https://graph.microsoft.com/v1.0{path}",
            headers=headers,
            content=content,
            json=json_payload,
        )


def _build_workbook_bytes(
    clinic_name: str,
    lead_tab_name: str,
    availability_enabled: bool,
    availability_tab_name: str,
) -> bytes:
    workbook = Workbook()
    lead_tab = lead_tab_name.strip() or "Leads"
    worksheet = workbook.active
    worksheet.title = lead_tab
    worksheet.append(LEAD_HEADERS)

    if availability_enabled:
        availability_sheet = workbook.create_sheet(title=availability_tab_name or "Availability")
        availability_sheet.append(AVAILABILITY_HEADERS)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def create_connected_workbook_for_clinic(
    *,
    access_token: str,
    clinic_name: str,
    lead_tab_name: str,
    availability_enabled: bool,
    availability_tab_name: str,
) -> dict:
    file_name = f"{clinic_name} Leads.xlsx"
    workbook_bytes = _build_workbook_bytes(
        clinic_name,
        lead_tab_name,
        availability_enabled,
        availability_tab_name,
    )
    encoded_name = quote(file_name)
    response = await _graph_request(
        "PUT",
        f"/me/drive/root:/{encoded_name}:/content",
        access_token,
        content=workbook_bytes,
    )
    payload = response.json()
    if response.status_code >= 400:
        detail = payload.get("error", {}).get("message") or "Microsoft workbook creation failed"
        raise RuntimeError(detail)

    return {
        "workbook_id": str(payload.get("id") or ""),
        "workbook_name": str(payload.get("name") or file_name),
        "workbook_url": str(payload.get("webUrl") or ""),
        "lead_tab_name": lead_tab_name.strip() or "Leads",
        "availability_tab_name": availability_tab_name or "Availability",
    }


async def _get_excel_access_token(clinic: dict) -> str:
    refresh_token = str(clinic.get("microsoft_excel_refresh_token") or "").strip()
    if not refresh_token:
        raise RuntimeError("Microsoft Excel is not fully connected for this clinic.")

    refreshed = await refresh_microsoft_access_token(refresh_token)
    access_token = str(refreshed.get("access_token") or "").strip()
    next_refresh_token = str(refreshed.get("refresh_token") or refresh_token).strip()
    if next_refresh_token and next_refresh_token != refresh_token:
        get_supabase().table("clinics").update(
            {"microsoft_excel_refresh_token": next_refresh_token}
        ).eq("id", clinic["id"]).execute()
    if not access_token:
        raise RuntimeError("Microsoft access token refresh did not return an access token.")
    return access_token


async def _download_workbook(clinic: dict) -> tuple[Workbook, str]:
    workbook_id = str(clinic.get("excel_workbook_id") or "").strip()
    if not workbook_id:
        raise RuntimeError("Microsoft Excel workbook is not configured for this clinic.")
    access_token = await _get_excel_access_token(clinic)
    response = await _graph_request("GET", f"/me/drive/items/{workbook_id}/content", access_token)
    if response.status_code >= 400:
        raise RuntimeError("Unable to fetch the connected Excel workbook.")
    workbook = load_workbook(BytesIO(response.content))
    return workbook, access_token


async def _upload_workbook(clinic: dict, workbook: Workbook, access_token: str) -> None:
    workbook_id = str(clinic.get("excel_workbook_id") or "").strip()
    output = BytesIO()
    workbook.save(output)
    response = await _graph_request(
        "PUT",
        f"/me/drive/items/{workbook_id}/content",
        access_token,
        content=output.getvalue(),
    )
    if response.status_code >= 400:
        raise RuntimeError("Unable to update the connected Excel workbook.")


async def append_lead_to_workbook(clinic: dict, lead: dict) -> None:
    workbook, access_token = await _download_workbook(clinic)
    tab_name = str(clinic.get("google_sheet_tab") or "Leads")
    if tab_name in workbook.sheetnames:
        worksheet = workbook[tab_name]
    else:
        worksheet = workbook.active
        worksheet.title = tab_name
        worksheet.append(LEAD_HEADERS)

    if worksheet.max_row == 1 and not worksheet.cell(1, 1).value:
        worksheet.append(LEAD_HEADERS)

    worksheet.append(
        [
            lead.get("id", ""),
            lead.get("created_at", ""),
            lead.get("patient_name", ""),
            lead.get("patient_phone", ""),
            lead.get("patient_email", ""),
            lead.get("reason_for_visit", ""),
            lead.get("preferred_datetime_text", ""),
            lead.get("status", "new"),
        ]
    )
    await _upload_workbook(clinic, workbook, access_token)


async def update_lead_status_in_workbook(clinic: dict, lead_id: str, new_status: str) -> None:
    workbook, access_token = await _download_workbook(clinic)
    tab_name = str(clinic.get("google_sheet_tab") or "Leads")
    if tab_name not in workbook.sheetnames:
        return
    worksheet = workbook[tab_name]
    for row in range(2, worksheet.max_row + 1):
        if str(worksheet.cell(row, 1).value or "") == lead_id:
            worksheet.cell(row, 8).value = new_status
            await _upload_workbook(clinic, workbook, access_token)
            return


async def get_available_slots_from_workbook(clinic: dict) -> list[dict]:
    workbook, _ = await _download_workbook(clinic)
    tab_name = str(clinic.get("availability_sheet_tab") or "Availability")
    if tab_name not in workbook.sheetnames:
        return []
    worksheet = workbook[tab_name]
    headers = [str(cell.value or "").strip() for cell in worksheet[1]]
    if not {"Date", "Time", "Status"}.issubset(set(headers)):
        return []
    slots: list[dict] = []
    for index in range(2, worksheet.max_row + 1):
        date = str(worksheet.cell(index, 1).value or "").strip()
        time_value = str(worksheet.cell(index, 2).value or "").strip()
        status = str(worksheet.cell(index, 3).value or "").strip().lower()
        if status == "available" and date and time_value:
            slots.append({"date": date, "time": time_value, "row_index": index})
    return slots


async def reserve_slot_in_workbook(
    clinic: dict,
    row_index: int,
    patient_name: str,
    lead_id: str,
) -> bool:
    workbook, access_token = await _download_workbook(clinic)
    tab_name = str(clinic.get("availability_sheet_tab") or "Availability")
    if tab_name not in workbook.sheetnames:
        return False
    worksheet = workbook[tab_name]
    worksheet.cell(row_index, 3).value = "reserved"
    worksheet.cell(row_index, 4).value = patient_name
    worksheet.cell(row_index, 5).value = lead_id
    await _upload_workbook(clinic, workbook, access_token)
    return True
